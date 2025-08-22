import os
import io
from tkinter import Tk, filedialog, messagebox, Label, Button, Entry, StringVar, ttk
from openpyxl import load_workbook
from PIL import Image, ImageFilter

def find_start_row(sheet, code_col_idx, family_col_idx):
    for row in range(1, sheet.max_row + 1):
        code_val = sheet.cell(row=row, column=code_col_idx).value
        family_val = sheet.cell(row=row, column=family_col_idx).value
        if code_val and family_val:
            return row
    return 1

def extract_images_from_excel(
    excel_path, output_dir, image_col, code_col, family_col, scale_factor=2.3, progress=None
):
    wb = load_workbook(excel_path)
    sheet = wb.active

    image_col_idx = ord(image_col.upper()) - ord('A') + 1
    code_col_idx = ord(code_col.upper()) - ord('A') + 1
    family_col_idx = ord(family_col.upper()) - ord('A') + 1

    start_row = find_start_row(sheet, code_col_idx, family_col_idx)
    total_images = len(sheet._images)

    for idx, image in enumerate(sheet._images, 1):
        row = max(image.anchor._from.row + 1, start_row)

        family_name = sheet.cell(row=row, column=family_col_idx).value
        item_code = sheet.cell(row=row, column=code_col_idx).value

        if family_name:
            family_name = str(family_name).replace("/", "").strip()
        if item_code:
            item_code = str(item_code).replace("/", "_").strip()

        if family_name and item_code:
            img_bytes = io.BytesIO(image._data())
            img = Image.open(img_bytes)

            new_size = (int(img.width * scale_factor), int(img.height * scale_factor))
            img_resized = img.resize(new_size, Image.LANCZOS)
            img_sharpened = img_resized.filter(ImageFilter.SHARPEN)

            output_path = os.path.join(output_dir, family_name)
            os.makedirs(output_path, exist_ok=True)

            img_path = os.path.join(output_path, f"{item_code}.jpg")
            img_sharpened.convert('RGB').save(img_path, format='JPEG', quality=95)

        if progress:
            progress['value'] = (idx / total_images) * 100
            progress.update()

def browse_file(entry):
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm")])
    entry.set(filename)

def browse_output_dir(entry):
    directory = filedialog.askdirectory()
    entry.set(directory)

def start_extraction():
    excel_file = excel_var.get()
    output_dir = output_dir_var.get()
    image_col = image_col_var.get()
    code_col = code_col_var.get()
    family_col = family_col_var.get()
    scale_factor = scale_factor_var.get()

    if not excel_file or not output_dir or not image_col or not code_col or not family_col or not scale_factor:
        messagebox.showerror("שגיאה", "חובה למלא את כל השדות")
        return

    try:
        scale = float(scale_factor)
    except ValueError:
        messagebox.showerror("שגיאה", "גודל התמונה חייב להיות מספר")
        return

    extract_images_from_excel(
        excel_path=excel_file,
        output_dir=output_dir,
        image_col=image_col,
        code_col=code_col,
        family_col=family_col,
        scale_factor=scale,
        progress=progress_bar
    )

    messagebox.showinfo("הצלחה", "התמונות חולצו בהצלחה לתיקייה המבוקשת")

# GUI setup
root = Tk()
root.title("חילוץ תמונות מאקסל לתיקיות משפחות")
root.geometry('520x480')

Label(root, text="קובץ Excel", font=("Arial", 12)).pack(pady=5)
excel_var = StringVar()
Entry(root, textvariable=excel_var, width=60).pack()
Button(root, text="בחר קובץ Excel", command=lambda: browse_file(excel_var)).pack(pady=5)

Label(root, text="תיקיית יעד לתמונות", font=("Arial", 12)).pack(pady=5)
output_dir_var = StringVar()
Entry(root, textvariable=output_dir_var, width=60).pack()
Button(root, text="בחר תיקייה", command=lambda: browse_output_dir(output_dir_var)).pack(pady=5)

Label(root, text="עמודת התמונות (למשל: B)").pack()
image_col_var = StringVar()
Entry(root, textvariable=image_col_var, width=10).pack()

Label(root, text="עמודת קוד (למשל: A)").pack()
code_col_var = StringVar()
Entry(root, textvariable=code_col_var, width=10).pack()

Label(root, text="עמודת משפחה (למשל: D)").pack()
family_col_var = StringVar()
Entry(root, textvariable=family_col_var, width=10).pack()

Label(root, text="גודל תמונה (ברירת מחדל: 2.3)").pack()
scale_factor_var = StringVar(value="2.3")
Entry(root, textvariable=scale_factor_var, width=10).pack()

progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress_bar.pack(pady=15)

Button(root, text="התחל חילוץ", command=start_extraction, font=("Arial", 14), bg="#4caf50", fg="white").pack(pady=10)

root.mainloop()
